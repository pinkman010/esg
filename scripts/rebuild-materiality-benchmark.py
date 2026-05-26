from __future__ import annotations

import json
import math
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

sys.path.insert(0, r"C:\tmp\codex_pydeps")

from openpyxl import Workbook, load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
DATASET_PATH = ROOT / "public" / "data" / "demo-dataset.json"
OUTPUT_DIR = ROOT / "outputs"
FULL_TOPIC_WORKBOOK = OUTPUT_DIR / "远景能源_全量分析实质性议题清单.xlsx"
REVIEW_OUTPUT = OUTPUT_DIR / "五家公司_实质性议题热力图_规则初筛.xlsx"

PARTIAL_WEIGHT = 0.6

TOPIC_ID_BY_NAME = {
    "商业道德与合规经营": "business-ethics-compliance",
    "社区关系与社会贡献": "community-relations-social-contribution",
    "产品质量与客户责任": "product-quality-customer-responsibility",
    "气候变化与温室气体排放": "climate-ghg",
    "能源使用与能效管理": "energy-efficiency",
    "资源使用与循环经济": "circular-resource",
    "创新研发与知识产权": "innovation-ip",
    "人权与价值链劳工权益": "human-rights-value-chain-workers",
    "供应链责任管理": "sustainable-supply-chain",
    "水资源管理": "water-management",
    "生物多样性与生态保护": "biodiversity",
    "废弃物管理": "waste-management",
    "数据安全与隐私保护": "data-security-privacy",
    "职业健康与安全": "employee-health-safety",
    "人才发展与员工培训": "talent-development-training",
    "公司治理与可持续发展治理": "esg-governance",
}

DIMENSION_MAP = {
    "环境": "E",
    "社会": "S",
    "治理": "G",
    "综合": "G",
}

REPORT_FILES = {
    "envision": "Envision Energy ESG Report 2024.pdf",
    "vestas": "Vestas Annual Report 2024.pdf",
    "siemens-energy": "se-sustainability-report-2024-pdf_Original file.pdf",
    "goldwind": "Goldwind_Sustainability_Report_2024_English.pdf",
    "mingyang": "MY-Sustainability Report_2024.pdf",
}

STOPWORDS = {
    "about",
    "across",
    "after",
    "also",
    "and",
    "another",
    "are",
    "based",
    "been",
    "being",
    "between",
    "both",
    "can",
    "company",
    "could",
    "data",
    "disclose",
    "disclosed",
    "disclosure",
    "does",
    "during",
    "each",
    "from",
    "have",
    "having",
    "how",
    "including",
    "into",
    "its",
    "main",
    "may",
    "more",
    "not",
    "other",
    "related",
    "report",
    "shall",
    "should",
    "such",
    "than",
    "that",
    "the",
    "their",
    "these",
    "this",
    "those",
    "through",
    "undertaking",
    "whether",
    "which",
    "with",
    "within",
    "year",
}


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value.replace("\u00a0", " ")).strip()


def tokenize(value: str) -> list[str]:
    return [
        token
        for token in re.findall(r"[a-z][a-z\-]{3,}", value.lower())
        if token not in STOPWORDS and not token.isdigit()
    ]


def split_keywords(value: str | None) -> list[str]:
    if not value:
        return []
    raw = re.split(r"[、,;；\s]+", value.lower())
    return [item.strip() for item in raw if len(item.strip()) >= 3]


def score_to_depth(score: int) -> str:
    if score >= 90:
        return "leading"
    if score >= 80:
        return "adequate"
    if score >= 70:
        return "weak"
    return "missing"


def read_pdf_pages(path: Path) -> list[str]:
    reader = PdfReader(str(path))
    if reader.is_encrypted:
        reader.decrypt("")
    pages: list[str] = []
    for page in reader.pages:
        try:
            pages.append(normalize_text(page.extract_text() or ""))
        except Exception:
            pages.append("")
    return pages


def load_topics() -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    wb = load_workbook(FULL_TOPIC_WORKBOOK, read_only=True, data_only=True)
    topic_sheet = wb["远景报告明确议题"]
    mapping_sheet = wb["全量分析议题映射"]

    topics: list[dict[str, Any]] = []
    for row in topic_sheet.iter_rows(min_row=2, values_only=True):
        topic_name = row[1]
        if not topic_name or topic_name not in TOPIC_ID_BY_NAME:
            continue
        total = int(row[8] or 0)
        disclosed = int(row[11] or 0)
        partial = int(row[12] or 0)
        missing = int(row[13] or 0)
        pending = int(row[14] or 0)
        score = round(((disclosed + partial * PARTIAL_WEIGHT) / total) * 100) if total else 0
        topics.append(
            {
                "topicId": TOPIC_ID_BY_NAME[topic_name],
                "topicName": topic_name,
                "dimension": DIMENSION_MAP.get(row[3], "G"),
                "keywords": split_keywords(row[20]),
                "envision": {
                    "total": total,
                    "disclosed": disclosed,
                    "partial": partial,
                    "missing": missing,
                    "pending": pending,
                    "score": score,
                    "pages": row[16] or "",
                    "evidence": row[17] or "",
                },
            }
        )

    requirements: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in mapping_sheet.iter_rows(min_row=2, values_only=True):
        topic_name = row[0]
        if topic_name not in TOPIC_ID_BY_NAME:
            continue
        requirement_text = normalize_text(str(row[8] or ""))
        tokens = Counter(tokenize(requirement_text)).most_common(18)
        requirements[topic_name].append(
            {
                "standard": row[2],
                "clause": row[3],
                "status": row[5],
                "gap": row[6],
                "priority": row[7],
                "text": requirement_text,
                "tokens": [token for token, _ in tokens],
            }
        )

    return topics, requirements


def classify_requirement(requirement: dict[str, Any], page_texts: list[str], topic_keywords: list[str]) -> tuple[str, int]:
    tokens = requirement["tokens"][:10]
    best_hits = 0
    best_page = 0
    for index, text in enumerate(page_texts):
        if not text:
            continue
        lower = text.lower()
        token_hits = sum(1 for token in tokens if token in lower)
        keyword_hits = sum(1 for keyword in topic_keywords if keyword in lower)
        hits = token_hits + min(keyword_hits, 4)
        if hits > best_hits:
            best_hits = hits
            best_page = index + 1

    if best_hits >= 7:
        return "disclosed", best_page
    if best_hits >= 4:
        return "partial", best_page
    if best_hits >= 2:
        return "pending", best_page
    return "missing", 0


def summarize_topic(company_id: str, topic: dict[str, Any], requirements: list[dict[str, Any]], page_texts: list[str]) -> dict[str, Any]:
    if company_id == "envision":
        counts = topic["envision"]
        score = int(counts["score"])
        pages = str(counts["pages"])
        evidence = (
            f"检索页码：{pages}。中文摘要：远景全量议题清单显示，“{topic['topicName']}”相关披露要求共{counts['total']}项，"
            f"其中已披露{counts['disclosed']}项、部分披露{counts['partial']}项、未披露{counts['missing']}项、"
            f"待确认{counts['pending']}项；综合得分{score}分。"
        )
        return {
            "score": score,
            "depth": score_to_depth(score),
            "total": counts["total"],
            "disclosed": counts["disclosed"],
            "partial": counts["partial"],
            "missing": counts["missing"],
            "pending": counts["pending"],
            "pages": pages,
            "evidence": evidence,
            "signal": f"远景全量规则初筛：已披露{counts['disclosed']}项、部分披露{counts['partial']}项、待确认{counts['pending']}项。",
        }

    status_counts: Counter[str] = Counter()
    page_counts: Counter[int] = Counter()
    for requirement in requirements:
        status, page = classify_requirement(requirement, page_texts, topic["keywords"])
        status_counts[status] += 1
        if page:
            page_counts[page] += 1

    total = len(requirements)
    disclosed = status_counts["disclosed"]
    partial = status_counts["partial"]
    missing = status_counts["missing"]
    pending = status_counts["pending"]
    score = round(((disclosed + partial * PARTIAL_WEIGHT) / total) * 100) if total else 0

    top_pages = [page for page, _ in page_counts.most_common(3)]
    if top_pages:
        pages = "、".join(f"P{page}" for page in top_pages)
        evidence = (
            f"检索页码：{pages}。中文摘要：报告在上述页码检索到与“{topic['topicName']}”相关的披露线索；"
            f"规则初筛识别相关披露要求共{total}项，其中已披露{disclosed}项、部分披露{partial}项、"
            f"未披露{missing}项、待确认{pending}项；综合得分{score}分，需结合原文页码人工复核。"
        )
    else:
        pages = "无"
        evidence = (
            f"未检索到“{topic['topicName']}”的可用证据。规则初筛识别相关披露要求共{total}项，"
            f"已披露{disclosed}项、部分披露{partial}项、未披露{missing}项、待确认{pending}项；综合得分{score}分。"
        )

    return {
        "score": int(score),
        "depth": score_to_depth(int(score)),
        "total": total,
        "disclosed": disclosed,
        "partial": partial,
        "missing": missing,
        "pending": pending,
        "pages": pages,
        "evidence": evidence,
        "signal": f"规则初筛：已披露{disclosed}项、部分披露{partial}项、未披露{missing}项、待确认{pending}项，需人工复核。",
    }


def rebuild_dataset(records: list[dict[str, Any]]) -> None:
    data = json.loads(DATASET_PATH.read_text(encoding="utf-8"))
    data["materialityBenchmark"] = [
        {
            "id": f"bench-{record['topicId']}-{record['companyId']}",
            "topicId": record["topicId"],
            "topicName": record["topicName"],
            "dimension": record["dimension"],
            "companyId": record["companyId"],
            "companyName": record["companyName"],
            "score": record["score"],
            "disclosureDepth": record["depth"],
            "evidence": record["evidence"],
            "signal": record["signal"],
            "sourceReport": record["sourceReport"],
        }
        for record in records
    ]
    DATASET_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_review_workbook(records: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "热力图规则初筛"
    headers = [
        "公司",
        "议题",
        "维度",
        "分数",
        "等级",
        "相关要求",
        "已披露",
        "部分披露",
        "未披露",
        "待确认",
        "页码",
        "证据摘要",
        "规则说明",
        "来源报告",
    ]
    sheet.append(headers)
    for record in records:
        sheet.append(
            [
                record["companyName"],
                record["topicName"],
                record["dimension"],
                record["score"],
                record["depth"],
                record["total"],
                record["disclosed"],
                record["partial"],
                record["missing"],
                record["pending"],
                record["pages"],
                record["evidence"],
                record["signal"],
                record["sourceReport"],
            ]
        )
    workbook.save(REVIEW_OUTPUT)


def main() -> None:
    topics, requirements_by_topic = load_topics()
    data = json.loads(DATASET_PATH.read_text(encoding="utf-8"))
    companies = data["companies"]

    page_cache: dict[str, list[str]] = {}
    for company in companies:
        report_file = REPORT_FILES[company["id"]]
        page_cache[company["id"]] = read_pdf_pages(ROOT / "data" / report_file)

    records: list[dict[str, Any]] = []
    for topic in topics:
        requirements = requirements_by_topic[topic["topicName"]]
        for company in companies:
            report_file = REPORT_FILES[company["id"]]
            result = summarize_topic(company["id"], topic, requirements, page_cache[company["id"]])
            records.append(
                {
                    **result,
                    "companyId": company["id"],
                    "companyName": company["name"],
                    "topicId": topic["topicId"],
                    "topicName": topic["topicName"],
                    "dimension": topic["dimension"],
                    "sourceReport": report_file,
                }
            )

    expected = len(topics) * len(companies)
    if len(records) != expected:
        raise RuntimeError(f"record count mismatch: {len(records)} != {expected}")

    rebuild_dataset(records)
    write_review_workbook(records)
    print(f"updated {DATASET_PATH}")
    print(f"wrote {REVIEW_OUTPUT}")
    print(f"topics={len(topics)} companies={len(companies)} records={len(records)}")


if __name__ == "__main__":
    main()
